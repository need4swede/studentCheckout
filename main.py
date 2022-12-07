
##### SETUP ##################

## 1. Change log directories | First line is for Mac, Second is for Windows
_log_directory = ["/Volumes/StudentLogs", r"\\192.168.49.10\StudentLogs"]

## 2. Set school domain
_domain_name = '@domain.edu'

######## IMPORTS ##############################
if 'imports':
    import barcode, datetime, labels, linecache, multiprocessing, os, shutil, subprocess, sys
    from datetime import datetime
    from barcode.writer import ImageWriter
    from n4s import fs, strgs, term, web
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, DEFAULT_FONT
    from reportlab.graphics import shapes
    from reportlab.graphics.barcode import createBarcodeDrawing
    from PyQt6 import QtCore
    from PyQt6.QtCore import Qt, QDir
    from PyQt6.QtGui import QIcon, QCursor, QFont, QShortcut, QKeySequence
    from PyQt6.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QMessageBox, QVBoxLayout, QHBoxLayout, QCheckBox

######## GLOBAL VARIABLES #####################
if 'global_variables':
    _app_name = 'Student Checkout Tool'       # APP NAME
    _student_id = ''                          # STUDENT ID NUM
    _student_name = ''                        # STUDENT NAME
    _student_fname = ''                       # STUDENT FNAME
    _student_lname = ''                       # STUDENT LNAME
    _student_email = ''                       # STUDENT EMAIL
    _student_alias = ''                       # STUDENT FNAME[0]LNAME[:1]
    _student_dir_name = ''                    # STUDENT DIRECTORY NAME
    _device_status = ''                       # CHECK-IN || CHECK-OUT
    _device_status_from_search = False        # CHECK FROM SEARCH
    _auto_generate_barcode = False            # PASS IN STUDENT INFO
    _logbook = ''                             # STUDENT LOGBOOK
    _network = False                          # CHECK NETWORK
    _user = f'{QDir.homePath()}'              # USER DIR
    _screen_width = ''                        # SCREEN WIDTH
    _screen_height = ''                       # SCREEN HEIGHT
    test = ''

## CHECK IN / OUT DEVICES
class DeviceStatus(QWidget):
    """
    Device Check-In / Check-Out Window
    """
    def __init__(self, status, email='', dir=''):
        super().__init__()

        ## IMPORT GLOBAL VARIABLES
        global _logbook, _student_fname, _student_lname, _student_id, _student_email, _device_status_from_search

        ## DIRECTORY
        if fs.system('is-mac'):
            _student_dir_name = f"{fs.root('user')}/Documents/Student Checkout Tool/students"
        else:
            _student_dir_name = f"{fs.root('user')}\Documents\Student Checkout Tool\students"

        ## LOAD LOGBOOK
        self.logbook = load_workbook(_logbook)
        self.logbook_sheet = self.logbook['Logbook']

        ## LOGBOOK HEADERS
        self.logbook_headers = ["First Name", 
        "Last Name", "Student ID", "Email Address", "Asset Tags", "Checked-Out", "Checked-In"]

        ## CREATE LAYOUT
        self.layout = QVBoxLayout()
        self.setFixedSize(250, 180)

        ## KEEP WINDOW ON TOP
        self.setWindowFlags(QtCore.Qt.WindowType.WindowStaysOnTopHint)

        ## PROMPT FOR STUDENT EMAIL
        self.student_email = QLineEdit()
        self.student_email.setPlaceholderText('Student Email...')

        ## CHECK-IN LAYOUT
        if status == "in":

            ## WINDOW TITLE
            self.setWindowTitle('Device Check-In')

            ## BUTTON
            self.update_status_btn = QPushButton('Check-In')
            self.update_status_btn.clicked.connect(self.check_in)
            self.student_email.returnPressed.connect(self.check_in)
        
        ## CHECK-OUT LAYOUT
        if status == "out":

            ## WINDOW TITLE
            self.setWindowTitle('Device Check-Out')

            ## BUTTON
            self.update_status_btn = QPushButton('Check-Out')
            self.update_status_btn.clicked.connect(self.check_out)
            
            ## PROMPT FOR ASSET TAG
            self.device_tag = QLineEdit()
            self.device_tag.setPlaceholderText('Asset Tag...')
            self.device_tag.returnPressed.connect(self.check_out)
            self.layout.addWidget(self.device_tag)
        
        ## AUTOMATICALLY ADD EMAIL FROM STUDENT INFO FILE
        if not email == '':
            self.student_email.setText(email)
            if fs.system('is-mac'):
                self.student_info_file = f"{_student_dir_name}/{dir}/student_info.txt"
            else:
                self.student_info_file = f"{_student_dir_name}\{dir}\student_info.txt"
        else:
            self.student_info_file = 'None'
        
        ## MANUAL INPUT OF STUDENT INFO
        if not _device_status_from_search:

            ## FIRST NAME
            self.student_fname = QLineEdit()
            self.layout.addWidget(self.student_fname)
            self.student_fname.setPlaceholderText('First Name...')

            ## LAST NAME
            self.student_lname = QLineEdit()
            self.layout.addWidget(self.student_lname)
            self.student_lname.setPlaceholderText('Last Name...')

            ## STUDENT ID
            self.student_id = QLineEdit()
            self.layout.addWidget(self.student_id)
            self.student_id.setPlaceholderText('Student ID...')
            if status == 'in':
                self.student_id.returnPressed.connect(self.check_in)
            if status == 'out':
                self.student_id.returnPressed.connect(self.check_out)

            ## TRY GLOBAL FIRST NAME
            if not _student_fname == '':
                self.student_fname.setText(_student_fname)
            
            ## TRY GLOBAL LAST NAME
            if not _student_lname == '':
                self.student_lname.setText(_student_lname)
            
            ## TRY GLOBAL STUDENT ID
            if not _student_id == '':
                self.student_id.setText(_student_id)
            
            ## TRY GLOBAL STUDENT EMAIL
            if not _student_email == '':
                self.student_email.setText(_student_email)

        else:
            self.layout.addWidget(self.student_email)

        ## AMEND LAYOUT
        self.layout.addWidget(self.update_status_btn)
        self.setLayout(self.layout)

        ## SHOW WINDOW
        self.show()

    ## CHECK-IN A DEVICE
    def check_in(self):

        ## IMPORT GLOBAL VARIABLES
        global _logbook

        ## IF WINDOW IS HIDDEN, DISPLAY IT
        if self.isHidden():
            self.show()

        ## SET RUN PARAMETER
        check_in = True

        ## CHECK FOR EMPTY FIELDS
        try:
            if self.student_fname.text() == '' or self.student_lname.text() == '' or self.student_id.text() == '':
                self.student_fname.setPlaceholderText("First Name...[REQUIRED]")
                self.student_lname.setPlaceholderText("Last Name...[REQUIRED]")
                self.student_id.setPlaceholderText("Student ID...[REQUIRED]")
                return
        except AttributeError:
            pass
        
        ## LOADS THE LOGBOOK
        logbook = self.logbook.active

        ## CHECK IF LOGBOOK IS OPEN
        try:
            self.logbook.save(_logbook)
        except PermissionError:
            self.hide()
            MainWindow.information_window('open_logbook_changes')
            check_in = False
            self.check_in()

        ## GET CURRENT DATE
        date = datetime.now().strftime("%B %d, %Y")

        ## STUDENT EMAIL
        if not self.student_email.text() == '':
            student_email = self.student_email.text().strip().lower()
        else:
            student_email = f"{self.student_fname.text().lower().strip()}.{self.student_id.text().strip()}{_domain_name}"

        ## WRITE CHANGES TO LOGBOOK
        if check_in:

            ## FIND ROW WITH STUDENT EMAIL
            for x in range(1, logbook.max_row+1):

                ## AVOID EMPTY CELLS
                if logbook.cell(row=x, column=4).value is not None:

                    ## IF EMAIL MATCHES
                    if logbook.cell(row=x, column=4).value.strip().lower() == student_email:

                        ## ADD CHECK-IN TIME
                        logbook.cell(row=x, column=7).value = date

            ## SAVE LOGBOOK CHANGES
            self.logbook.save(_logbook)

            ## WRITE CHANGES TO SERVER LOG
            self.server('write')

            ## CLOSE WINDOW
            self.close()

    ## CHECK-OUT A DEVICE
    def check_out(self):

        ## IMPORT GLOBAL VARIABLES
        global _device_status_from_search, _student_fname, _student_lname, _student_email, _student_id

        ## IF WINDOW IS HIDDEN, DISPLAY IT
        if self.isHidden():
            self.show()

        ## SET RUN PARAMETER
        check_out = True

        ## CHECK FOR EMPTY FIELDS
        try:
            if self.student_fname.text() == '' or self.student_lname.text() == '' or self.device_tag.text() == '' or self.student_id.text() == '':
                self.student_fname.setPlaceholderText("First Name...[REQUIRED]")
                self.student_lname.setPlaceholderText("Last Name...[REQUIRED]")
                self.device_tag.setPlaceholderText("Asset Tag...[REQUIRED]")
                self.student_id.setPlaceholderText("Student ID...[REQUIRED]")
                return
        except AttributeError:
            pass
        
        ## LOADS THE LOGBOOK
        logbook = self.logbook.active

        ## CHECK IF LOGBOOK IS OPEN
        try:
            self.logbook.save(_logbook)
        except PermissionError:
            self.hide()
            MainWindow.information_window('open_logbook_changes')
            check_out = False
            self.check_out()

        ## GET FIRST EMPTY ROW
        curr_row = logbook.max_row + 1

        ## GET CURRENT DATE
        date = datetime.now().strftime("%B %d, %Y")

        ## CLEAR LINECACHE
        linecache.clearcache()

        ## IF CHECK OUT FROM A SEARCH CALL
        if _device_status_from_search:

            ## READ STUDENT INFO TEXT
            fname = strgs.clean_text(Input=linecache.getline(self.student_info_file, 1), Casing='title')
            lname = strgs.clean_text(Input=linecache.getline(self.student_info_file, 2), Casing='title')
            id = linecache.getline(self.student_info_file, 3)
            email = linecache.getline(self.student_info_file, 4)
        
        ## IF CHECK OUT VIA MANUAL ENTRY
        else:

            ## READ STUDENT INFO ENTRY
            fname = strgs.clean_text(Input=self.student_fname.text(), Casing='title')
            lname = strgs.clean_text(Input=self.student_lname.text(), Casing='title')
            id = self.student_id.text()
            if _student_email == '':
                email = f"{self.student_fname.text().lower().strip()}.{self.student_id.text().strip()}{_domain_name}"
            else:
                email = _student_email

        ## WRITE TO LOGBOOK
        if check_out:
            logbook.cell(row=curr_row, column=1).value = fname
            logbook.cell(row=curr_row, column=2).value = lname
            logbook.cell(row=curr_row, column=3).value = id
            logbook.cell(row=curr_row, column=4).value = email
            logbook.cell(row=curr_row, column=5).value = self.device_tag.text().upper()
            logbook.cell(row=curr_row, column=6).value = date

            ## SAVE LOGBOOK CHANGES
            self.logbook.save(_logbook)

            ## WRITE CHANGES TO SERVER LOG
            self.server('write')

            ## CLOSE WINDOW
            self.close()

    ## BACKUP LOGBOOK
    def server(self, action=''):

        ## VERIFY NETWORK CONNECTION
        if web.network_test():
            if action == 'write':
                try:
                    shutil.copy(MainWindow.student_logbook_file, MainWindow.student_logbook_file_network)
                except Exception:
                    if fs.system('is-mac'):
                        fs.path_exists(Path=f"{MainWindow.appProfile}/local", Make=True)
                        fs.copy_file(Source=MainWindow.student_logbook_file, Destination=f"{MainWindow.appProfile}/local/student_device_logbook.xlsx")
                    else:
                        fs.path_exists(Path=f"{MainWindow.appProfile}\local", Make=True)
                        fs.copy_file(Source=MainWindow.student_logbook_file, Destination=f"{MainWindow.appProfile}\local\student_device_logbook.xlsx")

## CREATE NEW STUDENT ENTRY
class NewStudentEntry(QWidget):
    """
    New Student Entry Window
    """
    
    ## INITIALIZE WINDOW
    def __init__(self):
        super().__init__()

        global _auto_generate_barcode, _student_fname, _student_lname, _student_id

        ## CREATE LAYOUT
        self.layout = QVBoxLayout()
        self.setFixedSize(250, 195)

        ## KEEP WINDOW ON TOP
        self.setWindowFlags(QtCore.Qt.WindowType.WindowStaysOnTopHint)

        ## WINDOW TITLE
        self.setWindowTitle('New Student Entry')
        
        ## LAYOUT ELEMENTS
        self.newStudent_first_name = QLineEdit()
        self.newStudent_first_name.setPlaceholderText('First Name...')
        self.newStudent_first_name.clear()
        if _auto_generate_barcode:
            self.newStudent_first_name.setText(_student_fname)

        self.newStudent_last_name = QLineEdit()
        self.newStudent_last_name.setPlaceholderText('Last Name...')
        self.newStudent_last_name.clear()
        if _auto_generate_barcode:
            self.newStudent_last_name.setText(_student_lname)

        self.newStudent_id = QLineEdit()
        self.newStudent_id.setPlaceholderText('ID Number...')
        self.newStudent_id.clear()
        if _auto_generate_barcode:
            self.newStudent_id.setText(_student_id)

        _auto_generate_barcode = False

        self.newStudent_print = QCheckBox('Print label upon creation')
        self.newStudent_print.setChecked(True)

        self.newStudent_add = QPushButton('Create Entry')
        self.newStudent_add.clicked.connect(self.create_barcodes)

        self.newStudent_manual_settings = QCheckBox('Manual Settings')
        self.newStudent_manual_settings.stateChanged.connect(self.enable_manual_settings)

        self.newStudent_manual_row = QLabel('Row')
        self.newStudent_manual_row_input = QLineEdit()
        self.newStudent_manual_row.hide()
        self.newStudent_manual_row_input.hide()

        self.newStudent_manual_col = QLabel('Column')
        self.newStudent_manual_col_input = QLineEdit()
        self.newStudent_manual_col.hide()
        self.newStudent_manual_col_input.hide()
        
        ## AMEND LAYOUT
        self.layout.addWidget(self.newStudent_first_name)
        self.layout.addWidget(self.newStudent_last_name)
        self.layout.addWidget(self.newStudent_id)
        self.layout.addWidget(self.newStudent_manual_settings)
        self.layout.addWidget(self.newStudent_print)
        self.layout.addWidget(self.newStudent_add)
        self.layout.addWidget(self.newStudent_manual_row)
        self.layout.addWidget(self.newStudent_manual_row_input)
        self.layout.addWidget(self.newStudent_manual_col)
        self.layout.addWidget(self.newStudent_manual_col_input)
        self.setLayout(self.layout)

        ## PROGRAM ENTER / RETURN KEY
        self.newStudent_id.returnPressed.connect(self.create_barcodes)
        self.newStudent_manual_row_input.returnPressed.connect(self.create_barcodes)
        self.newStudent_manual_col_input.returnPressed.connect(self.create_barcodes)

        ## SHOW WINDOW
        self.show()

    ## CREATES STUDENT BARCODES
    def create_barcodes(self):

        ## IMPORT GLOBAL VARIABLES
        global _student_name, _student_id, _student_alias, _student_dir_name, _student_fname, _student_lname, _student_email

        ## CREATE STUDENT ALIAS
        new_student_alias = f"{self.newStudent_first_name.text()[0].lower()}{self.newStudent_last_name.text().lower()}"
        _student_alias = new_student_alias

        ## CREATE NEW STUDENT EMAIL
        new_student_email = f"{self.newStudent_first_name.text().lower()}.{self.newStudent_id.text()}{_domain_name}"

        ## CREATE NEW STUDENT DIRECTORY
        if fs.system('is-mac'):

            ## MAKE STUDENTS DIR
            fs.path_exists(Path=f"{fs.root('user')}/Documents/Student Checkout Tool/students", Make=True)

            ## MAKE NEW STUDENT DIR
            new_student_dir = f"{fs.root('user')}/Documents/Student Checkout Tool/students/{new_student_email.replace('.', '_').replace('@', '-')}"
            fs.path_exists(Path=new_student_dir, Make=True)

            ## MAKE NEW STUDENT ID LABEL
            self.new_student_id_label = f"{new_student_dir}/IDCardLabel/label.pdf"
            fs.path_exists(Path=f"{new_student_dir}/IDCardLabel", Make=True)
        else:

            ## MAKE STUDENTS DIR
            fs.path_exists(Path=f"{fs.root('user')}\Documents\Student Checkout Tool\students", Make=True)

            ## MAKE NEW STUDENT DIR
            new_student_dir = f"{fs.root('user')}\Documents\Student Checkout Tool\students\{new_student_email.replace('.', '_').replace('@', '-')}"
            fs.path_exists(Path=new_student_dir, Make=True)

            ## MAKE NEW STUDENT ID LABEL
            self.new_student_id_label = fr"{new_student_dir}\IDCardLabel\label.pdf"
            fs.path_exists(Path=f"{new_student_dir}\IDCardLabel", Make=True)

        ## GENERATE STUDENT INFO TXT
        with open(f"{new_student_dir}/student_info.txt", "a+") as studentInfoText:
            studentInfoText.write(self.newStudent_first_name.text() + "\n")
            studentInfoText.write(self.newStudent_last_name.text() + "\n")
            studentInfoText.write(self.newStudent_id.text() + "\n")
            studentInfoText.write(new_student_email + "\n")
        
        ## UPDATE GLOBAL VARIABLES
        _student_fname = strgs.clean_text(Input=self.newStudent_first_name.text().strip(), Casing='title')
        _student_lname = strgs.clean_text(Input=self.newStudent_last_name.text().strip(), Casing='title')
        _student_id = self.newStudent_id.text().strip()
        _student_email = new_student_email.strip()
        
        ## INITIALIZE IMAGE WRITER AND SET FONT STYLE
        image_writer = ImageWriter()
        if fs.system('is-mac'):
            image_writer.font_path = os.path.join("fonts", f"{fs.root('user')}/Programming/Git/py-barcodes/fonts/DejaVuSansMono.ttf")
            self.font_path = os.path.join("fonts", f"{fs.root('user')}/Programming/Git/py-barcodes/fonts/DejaVuSansMono.ttf")
        else:
            image_writer.font_path = os.path.join("fonts", MainWindow.server_paths[1] + r"\StudentCheckoutTool\fonts\DejaVuSansMono.ttf")
            self.font_path = os.path.join("fonts", MainWindow.server_paths[1] + r"\StudentCheckoutTool\fonts\DejaVuSansMono.ttf")
        
        ## CREATE STUDENT ID BARCODE
        create_barcode = barcode.get('code128', self.newStudent_id.text(), writer=image_writer)
        create_barcode.save(f"{new_student_dir}/student_id")

        ## CREATE STUDENT NAME BARCODE
        create_barcode = barcode.get('code128', new_student_alias, writer=image_writer)
        create_barcode.save(f"{new_student_dir}/student_name")

        ## CREATE STUDENT EMAIL BARCODE
        create_barcode = barcode.get('code128', new_student_email, writer=image_writer)
        create_barcode.save(f"{new_student_dir}/student_email")
        
        ## CREATE PDF OF BARCODE
        self.create_labels(new_student_email, 'avery_5160')

        ## UPLOAD ENTRY TO SERVER
        if fs.path_exists(MainWindow.student_dir_network):
            MainWindow.server(action='upload_student', input_path=new_student_dir)
        
        ## OPENS PDF ON COMPLETION OR NO
        if self.newStudent_print.isChecked():
            if fs.system('is-mac'):
                subprocess.Popen(["open", self.new_student_id_label])
            else:
                subprocess.Popen(self.new_student_id_label, shell=True)
        else:
            ## DISPLAY MESSAGE
            QMessageBox.information(self, "New Student", f"\nSuccessfully added student!\n\n{new_student_email}")

        ## CLOSE WINDOW
        NewStudentEntry.close(self)

    ## DRAWS TEXT AND BARCODE FOR PDF
    def draw_labels(self, label, width, height, obj):
        (labelstr, barcodestr) = obj
        label.add(shapes.String(18, 25, labelstr, fontName="Helvetica", fontSize=9))
        label.add(createBarcodeDrawing('Code128', value=barcodestr, width=190, height=20))

    ## LABEL SPECIFICATIONS
    def spec_labels(self, type):
        specs = []
        if type == 'avery_5160':
            # Brand	Number	label across	label down	label width	label height	label padding left	label padding top	paper width	paper height	paper left margin	paper top margin
            # Avery	5160	3	            10	        2.625 in	1 in	        0.125 in	        0 in	            8.5 in	    11 in	        0.19 in	            0.5 in
            specs = [215.9, # Paper Height
            279.4, # Paper Width
            3, # Horizontal count
            10, # Vertical count
            64, # Label Width
            25.4, # Label Height
            2, # Border radius
            0, # Left Margin
            1, # Right Margin
            7.5, # Top Margin
            0, # Left Padding
            1, # Right Padding
            1, # Top Padding
            7.5, # Bottom Badding
            1.90] # Row Gap
        return specs

    ## CREATES PDF AND INSERTS TEXT AND BARCODE
    def create_labels(self, barcode, type):

        ## CHECK IF MANUAL LABEL POSITIONING IS CHECKED
        manual_settings = self.newStudent_manual_settings.isChecked()

        ## CLEAR LINE CACHE
        linecache.clearcache()

        ## IF COUNTER FILE DOESN'T EXIST
        if not fs.path_exists(f"{MainWindow.appProfile}/label_count.txt"):
            
            ## INITIALIZE LABEL COUNTER
            with open(f"{MainWindow.appProfile}/label_count.txt", "a+") as labelCountText:
                labelCountText.write("1" + "\n")
                labelCountText.write("0" + "\n")
                labelCountText.close()
        
        ## READ MANUAL LABEL POSITION
        if manual_settings:

            ## IF NO VALUES ADDED, WILL DEFAULT TO 1:1
            try:
                self.label_row = int(self.newStudent_manual_row_input.text())
                self.label_column = int(self.newStudent_manual_col_input.text())
            except ValueError:
                self.label_row = 1
                self.label_column = 1
            
            ## KEEP ROW AND COLUMN VALUES WITHIN MARGINS
            if self.label_row < 1:
                self.label_row = 1
            if self.label_row > 10:
                self.label_row = 10
            if self.label_column < 1:
                self.label_column = 1
            if self.label_column > 3:
                self.label_column = 3
        
        ## READ AUTO LABEL POSITION
        else:
            self.label_row = int(linecache.getline(f"{MainWindow.appProfile}/label_count.txt", 1))
            self.label_column = int(linecache.getline(f"{MainWindow.appProfile}/label_count.txt", 2))
        
        ## COUNT LABEL POSITION - MAX COLUMNS 3, ADD ROW ONLY AT MAX COLUMN COUNT
        if not manual_settings:
            if self.label_column == 3:
                self.label_row += 1
                self.label_column = 1
            else:
                self.label_column += 1

        ## READ CURRENT LABEL COUNT
        curr_row = linecache.getline(f"{MainWindow.appProfile}/label_count.txt", 1)
        curr_col = linecache.getline(f"{MainWindow.appProfile}/label_count.txt", 2)

        ## ADD LABEL POSITION
        with open(f"{MainWindow.appProfile}/label_count.txt", "r+") as labelCountText:
            if int(curr_row) == 10 and int(curr_col) == 2:
                labelCountText.write(str(1) + "\n")
                labelCountText.write(str(1) + "\n")
            else:
                labelCountText.write(str(self.label_row) + "\n")
                labelCountText.write(str(self.label_column) + "\n")

        ## GET LABEL SPECIFICATIONS
        label_spec = self.spec_labels(type)
        specs = labels.Specification(
            label_spec[0], label_spec[1], label_spec[2], label_spec[3],
            label_spec[4], label_spec[5], corner_radius=label_spec[6],
            left_margin=label_spec[7], right_margin=label_spec[8], 
            top_margin=label_spec[9], left_padding=label_spec[10], 
            right_padding=label_spec[11], top_padding=label_spec[12],
            bottom_padding=label_spec[13], row_gap=label_spec[14])

        ## CREATE PDF PAGE
        sheet = labels.Sheet(specs, self.draw_labels, border=False)

        ## CLEAR LINE CACHE
        linecache.clearcache()

        ## READ LABEL COUNT
        with open(f"{MainWindow.appProfile}/label_count.txt", "r") as labelCountText:
            new_row = linecache.getline(f"{MainWindow.appProfile}/label_count.txt", 1)
            new_col = linecache.getline(f"{MainWindow.appProfile}/label_count.txt", 2)

        ## CAST ROW / COLUMN INFO AS INT
        self.label_row = int(self.label_row)
        self.label_column = int(self.label_column)
        new_row = int(new_row)
        new_col = int(new_col)

        ## CALCULATE ROW / COLUMN SKIP
        if new_row == 1:
            ## PRINT FIRST LABEL WITHOUT ANY SKIPS
            if new_col == 1:
                pass
            ## KEEPS ROW AS 1, AND SKIPS COLUMNS
            elif new_col == 2:
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 2:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1, YOU MUST SKIP ROW 1 AND R1C1, R1C2, R1C3, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 3:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 4:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 5:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 6:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 7:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 8:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 9:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col-2), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-2), (10, 3)))
        elif new_row == 10:
            ## ROWS MUST HAVE ALL RESPECTIVE COLUMNS SKIPPED MANUALLY
            ## TO SKIP ALL OF ROW 1 AND ROW 2, YOU MUST SKIP ROWS AND R1C1, R1C2, R1C3, R2C1, R2C2, ETC.
            ## HERE COLUMNS ARE CALCULATED BASED ON THEIR CURRENT POSITION, WITH - / +
            if new_col == 1:
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col+2), (10, 3)))
                sheet.partial_page(1, ((new_row-9, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-9, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-9, new_col+2), (10, 3)))
            elif new_col == 2:
                sheet.partial_page(1, ((new_row-1, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-1, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-2, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-3, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-4, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-5, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-6, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-7, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-8, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row-9, new_col-1), (10, 3)))
                sheet.partial_page(1, ((new_row-9, new_col), (10, 3)))
                sheet.partial_page(1, ((new_row-9, new_col+1), (10, 3)))
                sheet.partial_page(1, ((new_row, new_col-1), (10, 3)))
            elif new_col == 3:
                sheet.partial_page(1, ((new_row-1, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-1, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-1, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row-2, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-2, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-2, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row-3, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-3, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-3, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row-4, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-4, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-4, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row-5, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-5, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-5, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row-6, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-6, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-6, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row-7, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-7, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-7, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row-9, new_col-2), (1, 1)))
                sheet.partial_page(1, ((new_row-9, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row-9, new_col), (1, 1)))
                sheet.partial_page(1, ((new_row, new_col-1), (1, 1)))
                sheet.partial_page(1, ((new_row, new_col-2), (1, 1)))
        
        ## ADD LABEL TO PDF
        sheet.add_label((barcode, barcode))

        ## EXPORT PDF
        sheet.save(self.new_student_id_label)

        ## LAST LABEL WARNING MESSAGE
        if int(new_row) == 10 and int(new_col) == 2:
            self.hide()
            MainWindow.information_window(message='last_label')

    ## SHOW MANUAL SETTINGS
    def enable_manual_settings(self, state):
        if state == Qt.CheckState.Checked.value:
            self.setFixedSize(250, 306)
            self.newStudent_manual_row.show()
            self.newStudent_manual_row_input.show()
            self.newStudent_manual_col.show()
            self.newStudent_manual_col_input.show()
        else:
            self.setFixedSize(250, 195)
            self.newStudent_manual_row.hide()
            self.newStudent_manual_row_input.hide()
            self.newStudent_manual_row_input.clear()
            self.newStudent_manual_col.hide()
            self.newStudent_manual_col_input.hide()
            self.newStudent_manual_col_input.clear()

## SEARCH FOR STUDENT ENTRY
class StudentSearch(QWidget):
    """
    Search Window for Finding Students
    """
    
    ## INITIALIZE WINDOW
    def __init__(self):
        super().__init__()

        ## IMPORT GLOBAL VARIABLES
        global _auto_generate_barcode, _student_fname, _student_lname, _student_id, _student_dir_name

        ## DIRECTORY
        try:
            if fs.system('is-mac'):
                self.dir_list = fs.read_dir(f"{fs.root('user')}/Documents/Student Checkout Tool/students", Output='dirs')
                _student_dir_name = f"{fs.root('user')}/Documents/Student Checkout Tool/students"
            else:
                self.dir_list = fs.read_dir(f"{fs.root('user')}\Documents\Student Checkout Tool\students", Output='dirs')
                _student_dir_name = f"{fs.root('user')}\Documents\Student Checkout Tool\students"
        except (TypeError, FileNotFoundError):
            pass

        ## LAYOUT
        self.setWindowTitle("Search")
        self.show()
        self.layout = QVBoxLayout() # MAIN WINDOW
        self.search_header = QHBoxLayout() # PODSEARCH LABEL
        search_section = QHBoxLayout() # SEARCH BOX / FIND
        self.search_results = QHBoxLayout() # SEARCH RESULT

        ## SEARCH BOX LABEL
        self.search_label = QLabel(f'{_app_name} - Search')
        self.search_header.addWidget(self.search_label)

        ## SEARCH BOX
        self.search_box = QLineEdit()
        self.search_box.setFixedWidth(250)
        self.search_box.setPlaceholderText('Student Email...')
        search_section.addWidget(self.search_box)

        ## FIND BUTTON
        self.searchBtn = QPushButton('Find Student')
        self.searchBtn.clicked.connect(lambda: self.run_search(self.search_box.text()))
        search_section.addWidget(self.searchBtn)

        # SEARCH RESULT
        self.result = QLabel('')
        self.result_btn_print = QPushButton('Print Label')
        self.result_btn_check_in = QPushButton('Check-In')
        self.result_btn_check_in.clicked.connect(self.device_status_in)
        self.result_btn_check_out = QPushButton('Check-Out')
        self.result_btn_check_out.clicked.connect(self.device_status_out)
        self.result_btn_delete = QPushButton('Delete')
        self.result_btn_delete.clicked.connect(lambda: MainWindow.information_window(message='delete_user', user=_student_email))
        self.result.hide(); self.result_btn_print.hide()
        self.result_btn_check_in.hide()
        self.result_btn_check_out.hide()
        self.result_btn_delete.hide()
        self.result_btn_print.setFixedWidth(80)

        ## CREATE - HEADER
        self.layout.addLayout(self.search_header)

        ## CREATE - SEARCH SECTION
        self.search_results.addWidget(self.result)
        self.search_results.addWidget(self.result_btn_print)
        self.search_results.addWidget(self.result_btn_check_in)
        self.search_results.addWidget(self.result_btn_check_out)
        self.search_results.addWidget(self.result_btn_delete)
        self.layout.addLayout(search_section)

        ## CREATE - MAIN LAYOUT
        self.setLayout(self.layout)

        ## KEEP WINDOW ON TOP
        self.setWindowFlags(QtCore.Qt.WindowType.WindowStaysOnTopHint)
        self.move(self.geometry().topLeft())

        ## SET FOCUS TO SEARCH BOX
        self.search_box.setFocus()

        ## KEYBOARD SHORTCUTS
            ## CLICK FIND BUTTON
        self.click_find = QShortcut(QKeySequence('Return'), self)
        self.click_find.activated.connect(self.searchBtn.click)
            ## CLOSE WINDOW
        self.shortcut_close = QShortcut(QKeySequence('Ctrl+m'), self)
        self.shortcut_close.activated.connect(lambda: self.close())

        ## IF OFFLINE
        _network = web.network_test()
        if not _network:
            self.searchBtn.setEnabled(False)
            self.search_box.setPlaceholderText('Network Offline!')

    ## DISPLAY SEARCH RESULTS
    def display_results(self, student_dir, search):

        ## DISPLAY USER SEARCH PARAMETER IN SEARCH BOX
        search_text = self.search_box.text()
        self.search_box.clear()
        self.search_box.setPlaceholderText(search_text)

        ## REMOVE SEARCH LABEL AND MOVE UP SEARCH BOX
        self.search_label.setText('')
        self.layout.removeItem(self.search_header)

        ## ADD SEARCH RESULT SECTIONS
        if not self.result.isVisible(): 
            self.layout.addLayout(self.search_results)
            self.setLayout(self.layout)

        # DISPLAY - RESULTS
        self.result.setText(search)
        self.result.show()

        # DISPLAY - BUTTONS
        self.result_btn_print.show()
        self.result_btn_print.clicked.connect(self.print_label)
        self.result_btn_check_in.show()
        self.result_btn_check_out.show()
        self.result_btn_delete.show()

    ## SEARCH FOR STUDENT ENTRY
    def run_search(self, search):

        ## IMPORT GLOBAL VARIABLE
        global _student_email, _student_dir_name

        ## ADD DOMAIN NAME IF NOT IN SEARCH
        if not "@" in str(search):
            search = f"{search}{_domain_name}"

        ## SET GLOBAL VARIABLE
        _student_email = search

        ## AMEND SEARCH TO FIT DIRECTORY NAMING SCHEME
        self.student_dir = search.replace('.', '_').replace('@', '-')

        ## SEARCH FOR STUDENT DIR
        try:

            ## STUDENT FOUND
            if self.student_dir.strip() in self.dir_list:
                self.display_results(self.student_dir, search)
                self.search_box.setPlaceholderText(f"{search}")
                if not self.result.isVisible():
                    self.result.show()
                    self.result_btn_print.show()
                    self.result_btn_check_in.show()
                    self.result_btn_check_out.show()
                    self.result_btn_delete.show()

            ## STUDENT NOT FOUND
            else:
                self.search_box.clear()
                self.search_box.setPlaceholderText('Nothing Found')
                if self.result.isVisible():
                    self.result.hide()
                    self.result_btn_print.hide()
                    self.result_btn_check_in.hide()
                    self.result_btn_check_out.hide()
                    self.result_btn_delete.hide()

        ## NO STUDENT ENTRIES FOUND
        except AttributeError:
            MainWindow.information_window('search_no_entries')
            self.close()

    ## PRINT A NEW LABEL OF FOUND STUDENT
    def print_label(self):

        ## IMPORT GLOBAL VARIABLES
        global _auto_generate_barcode, _student_fname, _student_lname, _student_id, _student_dir_name

        ## ENABLE AUTO LABEL GENERATION
        _auto_generate_barcode = True

        ## CLEAR LINECACHE
        linecache.clearcache()
        
        ## GET STUDENT INFO
        _student_fname = linecache.getline(f"{_student_dir_name}/{self.student_dir}/student_info.txt", 1).strip()
        _student_lname = linecache.getline(f"{_student_dir_name}/{self.student_dir}/student_info.txt", 2).strip()
        _student_id = linecache.getline(f"{_student_dir_name}/{self.student_dir}/student_info.txt", 3).strip()

        ## OPEN NEW STUDENT ENTRY
        MainWindow.new_student()

    ## DEVICE CHECK-IN
    def device_status_in(self):
        global _device_status, _device_status_from_search
        _device_status = 'check-in'
        _device_status_from_search = True
        self.device_status('in')

    ## DEVICE CHECK-OUT
    def device_status_out(self):
        global _device_status, _device_status_from_search
        _device_status = 'check-out'
        _device_status_from_search = True
        self.device_status('out')

    ## NEW STUDENT DATA
    def device_status(self, status):
        self.change_status = DeviceStatus(status, _student_email, self.student_dir)
        self.close()

## MAIN APPLICATION
class MainWindow(QWidget):

    ## INITIALIZE APPLICATION & GUI
    def __init__(self, *args, **kwargs):
        super(QWidget, self).__init__(*args, **kwargs)

        ################################################################################### GLOBAL FLAGS

        ## IMPORT GLOBAL VARIABLES
        global _screen_width, _screen_height, _logbook, _network
        
        ## GET SCREEN DIMENSIONS
        screen = QApplication.primaryScreen()
        rect = screen.availableGeometry()
        self.screen_width = rect.width()
        self.screen_height = rect.height()

        ## UPDATE SCREEN DIMENSIONS
        _screen_width = self.screen_width
        _screen_height = self.screen_height

        ## ON WINDOW CLOSE
        app.aboutToQuit.connect(lambda: self.quit(True))

        ################################################################################## APP DIRECTORY
        self.appDir = f'{_user}/{_app_name}'
        self.appRoot = f"{fs.root('apps')}/{_app_name}"
        self.appIcon = f"{self.appRoot}/Icon.icns"

        ## SERVER PATH
        self.server_paths = [_log_directory[0], _log_directory[1]]

        ## APP DOCUMENTS DIRECTORY
        if fs.system('is-mac'):
            self.appDocs = f"{fs.root('docs')}/{_app_name}"
            self.appDocs_network = self.server_paths[0] + r"/StudentCheckoutTool"
        else:
            self.appDocs = f"{fs.root('docs')}\{_app_name}"
            self.appDocs_network = self.server_paths[1] + r"\StudentCheckoutTool"
        fs.path_exists(Path=self.appDocs, Make=True)

        ## USER PROFILE
        if fs.system('is-mac'):
            self.appProfile = f"{fs.root('userlib')}/{strgs.clean_text(_app_name, Remove_Spaces=True)}/profile"
        else:
            self.appProfile = f"{fs.root('userlib')}\Roaming\{strgs.clean_text(_app_name, Remove_Spaces=True)}\profile"
        fs.path_exists(self.appProfile, True)

        ## STUDENT LOGBOOK FILE
        if fs.system('is-mac'):
            self.student_logbook_file = self.appProfile + "/student_device_logbook.xlsx"
            self.student_logbook_file_network = self.server_paths[0] + r"/StudentCheckoutTool/log/student_device_logbook.xlsx"
        else:
            self.student_logbook_file = self.appProfile + "\student_device_logbook.xlsx"
            self.student_logbook_file_network = self.server_paths[1] + r"\StudentCheckoutTool\log\student_device_logbook.xlsx"

        ## CREATE LOGBOOK FILE
        _logbook = self.student_logbook_file
        if not fs.path_exists(self.student_logbook_file):
            self.create_workbook(database='logbook')

        ## STUDENT ENTRIES
        if fs.system('is-mac'):
            self.student_dir = f"{fs.root('user')}/Documents/Student Checkout Tool/students"
            self.student_dir_network = self.server_paths[0] + r"/StudentCheckoutTool/students"
            if not fs.path_exists(self.student_dir):
                if fs.path_exists(f"{fs.root('user')}/Documents/Student Devices"):
                    shutil.copytree(self.student_dir_network, self.student_dir)
        else:
            self.student_dir = f"{fs.root('user')}\Documents\Student Checkout Tool\students"
            self.student_dir_network = self.server_paths[1] + r"\StudentCheckoutTool\students"

        #################################################################################### APP WINDOW
        self.setWindowTitle(_app_name)
        self.setWindowIcon(QIcon('icon.ico'))

        ################################################################################# PARENT LAYOUT

        # MAIN LAYOUT
        self.layout = QVBoxLayout()

        # HEIGHT:                 WIDTH:
        self.setFixedHeight(200); self.setFixedWidth(315)

        # MARGINS
        self.layout.setContentsMargins(15, 15, 15, 20)

        # SET LAYOUT
        self.setLayout(self.layout)

        ############################################################################## LAYOUT SECTIONS
    
        ########################################################## TOP SECTION
        self.studentEntriesSection = QHBoxLayout()

        ####################################################### MIDDLE SECTION
        self.deviceManagementSection = QHBoxLayout()
        self.deviceManagementButtons = QVBoxLayout()

        ########################################################## MESSAGE BOX
        self.message = QMessageBox()
        self.message.setCursor(QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.message.move(self.geometry().center())

        ##############################################################################################

        ###################################################################################### STUDENT ID
        self.studentIDBox = QLineEdit()
        self.studentIDBox.setFixedSize(140, 23.3)
        self.studentIDBox.setPlaceholderText('Student ID Number....')
        self.studentIDBox.hide()

        ################################################################################# STUDENT SEARCH
        self.findBtn = QPushButton('Search for Student')
        self.findBtn.setFixedSize(130,32)
        self.findBtn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.findBtn.clicked.connect(lambda: self.search('student'))

        ################################################################################## CREATE ENTRY
        self.createBarcodeBtn = QPushButton('New Student')
        self.createBarcodeBtn.setFixedSize(130,32)
        self.createBarcodeBtn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.createBarcodeBtn.clicked.connect(self.new_student)

        ## DEVICE MANAGEMENT LABEL
        self.deviceActionLabel = QLabel('  Device Management')

        ## CHECK-IN BUTTON
        self.deviceCheckIn = QPushButton('Check-In')
        self.deviceCheckIn.setFixedSize(130,32)
        self.deviceCheckIn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.deviceCheckIn.clicked.connect(self.device_status_in)

        ## CHECK-OUT BUTTON
        self.deviceCheckOut = QPushButton('Check-Out')
        self.deviceCheckOut.setFixedSize(130,32)
        self.deviceCheckOut.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.deviceCheckOut.clicked.connect(self.device_status_out)

        ## VIEW LOG BUTTON
        self.deviceLog = QPushButton('View Log')
        self.deviceLog.setFixedSize(130,32)
        self.deviceLog.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.deviceLog.clicked.connect(self.view_log)

        ##################################################################################### STUDENT SECTION
        self.studentEntriesSection.addWidget(self.studentIDBox)
        self.studentEntriesSection.addWidget(self.createBarcodeBtn)
        self.studentEntriesSection.addWidget(self.findBtn)

        ################################################################################ DEVICE SECTION
        self.deviceManagementButtons.addWidget(self.deviceActionLabel)
        self.deviceManagementButtons.addWidget(self.deviceCheckIn)
        self.deviceManagementButtons.addSpacing(15)
        self.deviceManagementButtons.addWidget(self.deviceCheckOut)
        self.deviceManagementButtons.addSpacing(15)
        self.deviceManagementButtons.addWidget(self.deviceLog)
        self.deviceManagementSection.addLayout(self.deviceManagementButtons)

        ################################################################################### CREATE LAYOUT

        ## CLEAR TERMINAL
        term.clear()

        ## CREATE LAYOUT
        self.layout.addLayout(self.studentEntriesSection)
        self.layout.addLayout(self.deviceManagementSection)

        ############################################################################### KEYBOARD SHORTCUT

        ## RUN STUDENT SEARCH
        self.shortcut_find = QShortcut(QKeySequence('Ctrl+s'), self)
        self.shortcut_find.activated.connect(lambda: self.search('student'))

        ## 'ENTER' TO CREATE BARCODE
        self.kb_search = QShortcut(QKeySequence('Return'), self)
        self.kb_search.activated.connect(self.createBarcodeBtn.click)

        ## CLOSE WINDOW
        self.shortcut_close = QShortcut(QKeySequence('Ctrl+m'), self)
        self.shortcut_close.activated.connect(lambda: self.close())

        ################################################################################### CHECK NETWORK

        ## RUN NETWORK CHECK
        _network = web.network_test()

        ################################################################################### DOWNLOAD DATABASE

        self.server('download_students')
        self.server('download_log')

    ## SEARCH FUNCTION
    def search(self, action):
        if action == 'student':
            self.run_search = StudentSearch()
            self.run_search.show()
            return

    ## CREATE STUDENT WORKBOOK
    def create_workbook(self, database):
        workbook = Workbook()
        if database == 'logbook':

            ## SET ACTIVE WORKBOOK
            Student_Device_Logbook = workbook.active
            Student_Device_Logbook.title = "Logbook"

            ## SET HEADER FONT SIZE
            fontStyle = Font(size = "14", bold=True, color="FFFFFF", shadow="000000")
            fontSize = Font(size = "16")
            
            ## FIRST NAME
            Student_Device_Logbook["A1"] = "First Name"
            Student_Device_Logbook["A1"].font = fontStyle
            Student_Device_Logbook["A1"].fill = PatternFill(fgColor="000000", fill_type="solid")
            Student_Device_Logbook.column_dimensions["A"].width = 28
            
            ## LAST NAME
            Student_Device_Logbook["B1"] = "Last Name"
            Student_Device_Logbook["B1"].font = fontStyle
            Student_Device_Logbook["B1"].fill = PatternFill(fgColor="000000", fill_type="solid")
            Student_Device_Logbook.column_dimensions["B"].width = 28
            
            ## STUDENT ID NUMBER
            Student_Device_Logbook["C1"] = "Student ID"
            Student_Device_Logbook["C1"].font = fontStyle
            Student_Device_Logbook["C1"].fill = PatternFill(fgColor="D76B0C", fill_type="solid")
            Student_Device_Logbook.column_dimensions["C"].width = 10
            
            ## STUDENT EMAIL ADDRESS
            Student_Device_Logbook["D1"] = "Email Address"
            Student_Device_Logbook["D1"].font = fontStyle
            Student_Device_Logbook["D1"].fill = PatternFill(fgColor="A70CD7", fill_type="solid")
            Student_Device_Logbook.column_dimensions["D"].width = 36
            
            ## ASSET TAG NUMBER OF DEVICE
            Student_Device_Logbook["E1"] = "Asset Tag"
            Student_Device_Logbook["E1"].font = fontStyle
            Student_Device_Logbook["E1"].fill = PatternFill(fgColor="0C34D7", fill_type="solid")
            Student_Device_Logbook.column_dimensions["E"].width = 14
            
            ## DATE OF CHECK-OUT
            Student_Device_Logbook["F1"] = "Checked-Out"
            Student_Device_Logbook["F1"].font = fontStyle
            Student_Device_Logbook["F1"].fill = PatternFill(fgColor="D70C49", fill_type="solid")
            Student_Device_Logbook.column_dimensions["F"].width = 22
            
            ## DATE OF CHECK-IN
            Student_Device_Logbook["G1"] = "Checked-In"
            Student_Device_Logbook["G1"].font = fontStyle
            Student_Device_Logbook["G1"].fill = PatternFill(fgColor="39AB10", fill_type="solid")
            Student_Device_Logbook.column_dimensions["G"].width = 22

            ## SET HEADER HEIGHT
            for x in range(1, 1000):
                Student_Device_Logbook.row_dimensions[x].height = 24

            ## SAVE FILE
            {k: setattr(DEFAULT_FONT, k, v) for k, v in fontSize.__dict__.items()}
            workbook.save(self.student_logbook_file)

    ## FIND STUDENT DATA
    def find_student(self, lookup):
        self.studentIDBox.clear()
        print(lookup)

    ## NEW STUDENT DATA
    def new_student(self):
        self.new_student_entry = NewStudentEntry()

    ## DELETE STUDENT DATA
    def remove_student(self):

        ## IMPORT GLOBAL VARIABLE
        global _student_email

        ## SET USER VARIABLE
        delete_user = _student_email

        ## AMEND SEARCH TO FIT DIRECTORY NAMING SCHEME
        delete_user = delete_user.replace('.', '_').replace('@', '-')

        ## GET USER DIRECTORY
        if fs.system('is-mac'):
            delete_user_dir = f"{self.student_dir}/{delete_user}"
        else:
            delete_user_dir = f"{self.student_dir}\{delete_user}"

        ## REMOVE USER ENTRY
        fs.remove_dir(delete_user_dir)

        ## CREATE NETWORK 'REMOVED' DIR AND SET MOVE DIR
        if fs.path_exists(self.appDocs_network):
            if fs.system('is-mac'):
                fs.path_exists(Path=f"{self.appDocs_network}/removed", Make=True)
                move_user_dir_source = f"{self.student_dir_network}/{delete_user}"
                move_user_dir_destination = f"{self.appDocs_network}/removed/{delete_user}"
            else:
                if not os.path.exists(f"{self.appDocs_network}\\removed"):
                    os.makedirs(f"{self.appDocs_network}\\removed")
                move_user_dir_source = f"{self.student_dir_network}\{delete_user}"
                move_user_dir_destination = f"{self.appDocs_network}\\removed\\{delete_user}"
        else:
            pass

        ## MOVE DELETED STUDENT TO 'REMOVED' NETWORK DIR
        try:
            shutil.copytree(move_user_dir_source, move_user_dir_destination, dirs_exist_ok=True)
            shutil.rmtree(move_user_dir_source)
        except Exception as e:
            print(e)
            pass

    ## DEVICE CHECK-IN
    def device_status_in(self):

        ## IMPORT GLOBAL VARIABLES
        global _device_status, _device_status_from_search

        ## SET GLOBAL CHECK-IN STATUS
        _device_status = 'check-in'

        ## SET GLOBAL SEARCH STATUS
        _device_status_from_search = False

        ## RUN DEVICE CHECK-IN
        self.device_status('in')

    ## DEVICE CHECK-OUT
    def device_status_out(self):

        ## IMPORT GLOBAL VARIABLES
        global _device_status, _device_status_from_search

        ## SET GLOBAL CHECK-IN STATUS
        _device_status = 'check-out'

        ## SET GLOBAL SEARCH STATUS
        _device_status_from_search = False

        ## RUN DEVICE CHECK-OUT
        self.device_status('out')

    ## NEW STUDENT DATA
    def device_status(self, status):
        self.change_status = DeviceStatus(status)

    ## SERVER TASKS
    def server(self, action='', input_path=''):

        ## VERIFY NETWORK CONECTION
        if web.network_test():

            ## DOWNLOAD LOGBOOK TO LOCAL DIRECTORY
            if action == 'download_log':
                try:
                    shutil.copy(self.student_logbook_file_network, self.student_logbook_file)
                except (FileExistsError, FileNotFoundError):
                    pass
                except (PermissionError):
                    self.information_window('open_logbook')
                    pass

            ## DOWNLOAD STUDENT ENTRIES
            if action == 'download_students':
                if fs.system('is-mac') and fs.path_exists(self.student_dir_network):
                    fs.remove_dir(self.student_dir)
                    shutil.copytree(self.student_dir_network, self.student_dir)

            ## UPLOAD STUDENT ENTRIES
            if action == 'upload_student':
                if fs.system('is-mac') and fs.path_exists(self.student_dir_network):
                    network_path = input_path.split("/")[-1]
                    shutil.copytree(input_path, f"{self.student_dir_network}/{network_path}", dirs_exist_ok=True)
                else:
                    network_path = input_path.split("\\")[-1]
                    shutil.copytree(input_path, f"{self.student_dir_network}\{network_path}", dirs_exist_ok=True)

    ## VIEW STUDENT DEVICE LOG
    def view_log(self):

        ## GET LOGBOOK FILE FROM SERVER
        if fs.path_exists(self.student_logbook_file_network):

            ## DOWNLOAD LOGBOOK FILE TO LOCAL DIRECTORY
            self.server('download_log')

        ## VIEW LOG ACTION
        if fs.system("is-mac"):

            ## VERIFY LOGBOOK LOCAL FILE PATH
            if fs.path_exists(self.student_logbook_file):

                ## VIEW LOGBOOK FILE
                subprocess.Popen(["open", (self.student_logbook_file)])

            else:

                ## DISPLAY MESSAGE
                self.information_window('no_logbook')

        else:

            ## VERIFY LOGBOOK LOCAL FILE PATH
            if fs.path_exists(self.student_logbook_file):

                ## VIEW LOGBOOK FILE
                try:
                    os.system(f"start EXCEL.EXE {self.student_logbook_file}")
                except PermissionError as e:
                    print(e)
                    self.information_window('open_logbook')
                    pass

            else:

                ## DISPLAY MESSAGE
                self.information_window('no_logbook')

    ## DISPLAY WARNING MESSAGE AT STARTUP
    def information_window(self, message, user=''):

        ## DELETING STUDENT ENTRY
        if message == 'delete_user':
            QMessageBox.move(self, int(self.screen_width/3), int(self.screen_height/3))
            prompt_delete_user = QMessageBox.question(self, "Warning", f'Are you sure you want to delete this entry?\n\n{user}\n\nYou can always add them again later',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.Yes)
            if prompt_delete_user == QMessageBox.StandardButton.Yes:
                self.remove_student()
                QMessageBox.information(self, "Removed Entry", 'Entry has been removed!')
                fs.system('python-restart')
            elif prompt_delete_user == QMessageBox.StandardButton.No:
                QMessageBox.information(self, "Removed Entry", 'Nothing removed')

        ## TRYING A SEARCH WITHOUT ANY STUDENT ENTRIES
        if message == 'search_no_entries':
            QMessageBox.move(self, int(self.screen_width/3), int(self.screen_height/3))
            QMessageBox.information(self, "Warning", 'No Student Entries Found!\n\nPlease add a new student first')

        ## CANNOT FIND LOGBOOK FILE
        if message == 'no_logbook':
            QMessageBox.move(self, int(self.screen_width/3), int(self.screen_height/3))
            QMessageBox.information(self, "Warning", 'No Student Device Log Found!')

        ## LOGBOOK IS OPEN (STARTUP)
        if message == 'open_logbook':
            QMessageBox.move(self, int(self.screen_width/3), int(self.screen_height/3))
            QMessageBox.information(self, "Warning", 'The Student Device Logbook is currently open.\nPlease close it to sync new entries.\n\nNote: This is the Excel spreadsheet that logs devices')

        ## LOGBOOK IS OPEN (NOT CURRENTLY USED FOR ANYTHING)
        if message == 'open_logbook_changes':
            QMessageBox.move(self, int(self.screen_width/3), int(self.screen_height/3))
            QMessageBox.information(self, "Warning", 'The Student Device Logbook is currently open.\nPlease close it and click "Ok".\n\nNote: This is the Excel spreadsheet that logs devices')

        ## DOWN TO THE LAST LABEL
        if message == 'last_label':
            QMessageBox.move(self, int(self.screen_width/3), int(self.screen_height/3))
            QMessageBox.information(self, "Warning", 'You are down to the last label!\n\nPlease replace the sheet for the next label')

    ## RESTART APP
    def restart(self):
        fs.system('python-restart')

    ## QUIT APP PROCESS
    def quit(self, clear: bool=False):
        app.exit()

## RUN
if __name__ == '__main__':

    ## REQUIRED FOR MULTI-THREADED WORKFLOW
    multiprocessing.freeze_support()

    ## INITIALIZE QAPP AND SET STYLESHEET
    app = QApplication(sys.argv)
    if fs.system('is-mac'):
        app.setStyleSheet('''
        * {
            background-color: #333;
        }
        QWidget {
            font-size: 15px;
            border-radius: 4px;
        }
        QLabel {
            font-family: 'Sans Serif';
        }
        QToolTip {
            padding: 4px; 
            border: 1px solid #bababa;
        }
        QStatusBar {
            font-size: 13px;
        }
        QStatusBar QPushButton {
            background-color: none;
            padding: 0 40px;
            color: #fff;
        }
        QStatusBar QPushButton:hover {
            background-color: none;
            color: #0078d4;
        }
        QLineEdit {
            padding: 4px 10px;
            margin-right: 10px;
            border: 2px solid #bababa;
            font-size: 16px;
            selection-background-color: #0078d4;
        }
        QLineEdit:hover {
            border-color: #808080;
        }
        QLineEdit:focus {
            border-color: #0078d4;
        }
        QMenu {
            border: 1px solid #bababa;
            padding: 5px;
        }
        QMenu::item {
            padding: 3px 25px;
            border-radius: 4px; 
        }
        QMenu::item:selected {
            color: #fff;
            background-color: #0078d4;
        }
        QPushButton {
            font-size: 12px;
            width: 0px;
            height: 10px;
            padding: 0;
            color: #fff;
            border: none;
            background-color: #656565;
        }
        QPushButton:hover, QComboBox:hover {
            background-color: #097ed9;
        }
        QPushButton:pressed, QComboBox:pressed {
            background-color: #00477c;
        }
        QPushButton:disabled, QComboBox:disabled {
            background-color: #77b7e9;
        }
        QComboBox {
            padding: 5.5px 30px 5.5px 45px;
            color: #fff;
            border: none;
            background-color: #0078d4;
        }
        QComboBox::drop-down {
            border-radius: 0;
        }
        QComboBox:on {
            border-bottom-left-radius: 0;
            border-bottom-right-radius: 0;
        }
        QComboBox QAbstractItemView {
            border-radius: 0;
            outline: 0;
        }
        QComboBox QAbstractItemView::item {
            height: 33px;
            padding-left: 42px;
            background-color: #fff;
        }
        QComboBox QAbstractItemView::item:selected {
            background-color: #0078d4;
        }
        QProgressBar {
            text-align: center;
        }
        QProgressBar::chunk {
            background: #0078d4;
            border-radius: 4px;
        }
        QMessageBox QLabel {
            font-size: 13px;
        }
        QMessageBox QPushButton {
            width: 60px;
            padding: 6px 8px;
        }
    ''')
        app.setFont(QFont('Helvetica Nue'))
        app.setStyleSheet("QLabel{font-family: 'Helvetica Nue';}")
    else:
        import qdarktheme
        app.setStyleSheet(qdarktheme.load_stylesheet())
    clipboard = app.clipboard()

    ## MAIN APPLICATION
    MainWindow = MainWindow()
    MainWindow.show()
    
    sys.exit(app.exec())